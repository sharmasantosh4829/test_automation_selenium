__maintainer__ = ['santosh.sharma']

import os
import logging
from openpyxl import workbook, load_workbook, worksheet


class Excl:
    path = os.getcwd()
    par_dir = os.path.abspath(os.path.join(path, os.pardir))
    testdata_dir = os.path.join(par_dir, "testdata")
    filename = __file__.split("/")[-1].split(".")[0]

    @staticmethod
    def get_row_count(file, sheetname):
        workbook = load_workbook(file)
        sheet = workbook[sheetname]
        return sheet.max_row

    @staticmethod
    def get_column_count(file, sheetname):
        workbook = load_workbook(file)
        sheet = workbook[sheetname]
        return sheet.max_column

    @staticmethod
    def read_data(file, sheetname, rownum, colnum):
        workbook = load_workbook(file)
        sheet = workbook[sheetname]
        return sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def write_data(file, sheetname, rownum, colnum, data):
        workbook = load_workbook(file)
        sheet = workbook[sheetname]
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(file)
